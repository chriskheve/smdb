# views.py
from openpyxl import load_workbook
from rest_framework.parsers import MultiPartParser, FormParser
from django.utils.text import slugify

from rest_framework.decorators import api_view, authentication_classes, permission_classes, parser_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.pagination import PageNumberPagination
from .serializers import SitesSerializer
from vendor.models import Vendor
from account.models import User
from risk_assessment.models import RiskAssessment
from .models import Sites
from rest_framework.pagination import PageNumberPagination
from django.db.models import Count
from django.db.models import Q
# import pandas as pd
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication

# from drf_spectacular.utils import extend_schema, OpenApiResponse, OpenApiExample, inline_serializer
from drf_spectacular.utils import (
    extend_schema, OpenApiParameter, OpenApiResponse, OpenApiExample, inline_serializer
)
from drf_spectacular.types import OpenApiTypes
from rest_framework import serializers

@extend_schema(
    tags=["sites"],
    summary="Créer un site",
    description="Crée un site avec `name`, `site_id`, coordonnées et relations (Vendor, RiskAssessment, ZM).",
    request=inline_serializer(
        name="CreateSiteRequest",
        fields={
            "name": serializers.CharField(help_text="Nom du site"),
            "site_id": serializers.CharField(help_text="Identifiant métier du site"),
            "latitude": serializers.FloatField(required=False, allow_null=True, help_text="Ex: -4.325"),
            "longitude": serializers.FloatField(required=False, allow_null=True, help_text="Ex: 15.322"),
            "vendor": serializers.UUIDField(required=False, allow_null=True, help_text="ID du Vendor"),
            "risk_assessment": serializers.UUIDField(required=False, allow_null=True, help_text="ID du RiskAssessment"),
            "zm": serializers.UUIDField(required=False, allow_null=True, help_text="ID de l'utilisateur ZM"),
        },
    ),
    responses={
        201: OpenApiResponse(response=SitesSerializer, description="Site créé"),
        400: OpenApiResponse(
            response=inline_serializer(name="SiteCreateBadRequest", fields={"message": serializers.CharField()}),
            description="Champs requis manquants / validation",
        ),
        404: OpenApiResponse(
            response=inline_serializer(name="SiteCreateNotFound", fields={"message": serializers.CharField()}),
            description="Ressource liée introuvable (Vendor / RiskAssessment / ZM)",
        ),
        500: OpenApiResponse(
            response=inline_serializer(name="SiteCreateServerError", fields={"message": serializers.CharField()}),
            description="Erreur interne",
        ),
    },
    examples=[
        OpenApiExample(
            "Requête complète",
            value={
                "name": "KASALA",
                "site_id": "CDKN00001",
                "latitude": -4.325,
                "longitude": 15.322,
                "vendor": "59a0f6f9-b6ec-4d5c-8a67-9f1e1d1b1c11",
                "risk_assessment": "3b4d0348-0c7c-4f18-9e0b-1b796b7b9d10",
                "zm": "c7b0c8b1-1c0e-4803-8d6e-3a1b799c9f22"
            },
            request_only=True,
        ),
        OpenApiExample(
            "Réponse 201 (exemple)",
            value={
                "id": "0b8c2b3e-9f3b-4a1e-bf4b-1f3f1d2c9a11",
                "name": "KASALA",
                "site_id": "CDKN00001",
                "latitude": -4.325,
                "longitude": 15.322,
                "vendor": "59a0f6f9-b6ec-4d5c-8a67-9f1e1d1b1c11",
                "risk_assessment": "3b4d0348-0c7c-4f18-9e0b-1b796b7b9d10",
                "zm": "c7b0c8b1-1c0e-4803-8d6e-3a1b799c9f22"
            },
            response_only=True,
        ),
        OpenApiExample(
            "Réponse 400",
            value={"message": "Les champs suivants sont requis : name, site_id"},
            response_only=True,
        ),
    ],
)

@api_view(['POST'])
@authentication_classes([])
@permission_classes([])
def create_site(request):
    """
    Crée un site avec des relations à Province, Battery, Vendor, Dg et User.
    """
    data = request.data


    # Validation des champs requis
    required_fields = ['name', 'site_id']
    missing_fields = [field for field in required_fields if field not in data]
    if missing_fields:
        return Response(
            {"message": f"Les champs suivants sont requis : {', '.join(missing_fields)}"},
            status=status.HTTP_400_BAD_REQUEST
        )
    vendor = None
    if 'vendor' in data:
        try:
            vendor = Vendor.objects.get(id=data['vendor'])
        except Vendor.DoesNotExist:
            return Response({"message": "Fournisseur introuvable."}, status=status.HTTP_404_NOT_FOUND)
        
    risk_assessment = None
    if 'risk_assessment' in data:
        try:
            risk_assessment = RiskAssessment.objects.get(id=data['risk_assessment'])
        except Vendor.DoesNotExist:
            return Response({"message": "Fournisseur introuvable."}, status=status.HTTP_404_NOT_FOUND)

    zm = None
    if 'zm' in data:
        try:
            zm = User.objects.get(id=data['zm'])
        except User.DoesNotExist:
            return Response({"message": "Utilisateur (ZM) introuvable."}, status=status.HTTP_404_NOT_FOUND)

    # Préparation des données pour la création du site
    site_data = {
        'name': data['name'],
        'site_id': data['site_id'],
        'latitude': data.get('latitude'),
        'longitude': data.get('longitude'),
        'vendor': vendor,
        'risk_assessment': risk_assessment,
        'zm': zm
    }

    try:
        # Création du site
        site = Sites.objects.create(**site_data)
        serialized_site = SitesSerializer(site)
        return Response(serialized_site.data, status=status.HTTP_201_CREATED)
    except Exception as e:
        print(e)
        return Response({"message": {str(e)}}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



def _norm(s: str) -> str:
    """Normalise un en-tête: 'EI Site ID' -> 'ei_site_id'."""
    return slugify(str(s).strip().lower()).replace('-', '_') if s is not None else ''
@extend_schema(
    tags=["sites"],
    summary="Importer des sites via Excel (.xlsx)",
    description=(
        "Importe des sites depuis un fichier Excel (.xlsx).\n\n"
        "**Colonnes attendues (insensibles à la casse/espaces)** :\n"
        "- `EI Site ID` → `site_id`\n"
        "- `Site Name` → `name`\n"
        "- `Latitude` → `latitude`\n"
        "- `Longitude` → `longitude`\n"
        "- `Vendor` → correspond à **Vendor.name** (recherche insensible à la casse)\n"
        "- `risk_assessment` → correspond à **RiskAssessment.name** (insensible à la casse)\n"
        "- `zm` → correspond à **User.id** (UUID)\n\n"
        "**Règles** :\n"
        "- Si `site_id` existe déjà, la ligne est ignorée (compteur `skipped`).\n"
        "- Les valeurs manquantes sur `name`/`site_id` font ignorer la ligne avec message dans `errors`.\n"
        "- Les Vendor/RiskAssessment/ZM inconnus n’empêchent pas la création du site : "
        "la ligne est créée sans le lien et un message est ajouté dans `errors`."
    ),
    request={
        "multipart/form-data": {
            "type": "object",
            "properties": {
                "file": {"type": "string", "format": "binary"}  # <-- bouton 'Choose file'
            },
            "required": ["file"]
        }
    },
    responses={
        200: OpenApiResponse(
            response=inline_serializer(
                name="SitesImportSummary",
                fields={
                    "created": serializers.IntegerField(),
                    "skipped": serializers.IntegerField(),
                    "errors": serializers.ListField(child=serializers.CharField()),
                },
            ),
            description="Résumé de l'import"
        ),
        400: OpenApiResponse(
            response=inline_serializer(name="ImportBadRequest", fields={"message": serializers.CharField()}),
            description="Requête invalide / fichier manquant / format incorrect"
        ),
        500: OpenApiResponse(
            response=inline_serializer(name="ImportServerError", fields={"message": serializers.CharField()}),
            description="Erreur interne"
        ),
    },
    examples=[
        OpenApiExample(
            "Entêtes attendues",
            value={"headers": ["EI Site ID", "Site Name", "Latitude", "Longitude", "Vendor", "risk_assessment", "zm", "security_type"]},
            response_only=True,
        ),
        OpenApiExample(
            "Réponse 200 (exemple)",
            value={"created": 42, "skipped": 3, "errors": ["L7: Vendor introuvable (name='Unknown')."]},
            response_only=True,
        ),
        OpenApiExample(
            "Réponse 400 (fichier manquant)",
            value={"message": "Aucun fichier reçu (champ 'file')."},
            response_only=True,
        ),
    ],
)
@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
@authentication_classes([])  # mets IsAuthenticated/JWT si tu veux protéger
@permission_classes([])
def import_sites_excel(request):
    f = request.FILES.get("file")
    if not f:
        return Response({"message": "Aucun fichier reçu (champ 'file')."}, status=400)
    if not f.name.lower().endswith(".xlsx"):
        return Response({"message": "Format non supporté. Utilise un .xlsx."}, status=400)

    try:
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return Response({"message": f"Lecture Excel impossible: {e}"}, status=400)

    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return Response({"message": "Fichier vide."}, status=400)

    # Normalisation des entêtes
    norm_headers = [_norm(h) for h in header]
    idx = {h: i for i, h in enumerate(norm_headers)}

    # Mapping des entêtes du fichier vers les champs Django
    # (EI Site ID -> site_id, Site Name -> name)
    required_map = {"ei_site_id": "site_id", "site_name": "name"}
    for needed in required_map.keys():
        if needed not in idx:
            return Response({"message": f"Colonne requise manquante: {needed}."}, status=400)

    created, skipped, errors = 0, 0, []

    for rnum, row in enumerate(rows, start=2):
        try:
            def get_col(key_norm):
                i = idx.get(key_norm)
                if i is None:
                    return None
                val = row[i]
                return None if val in (None, "") else val

            site_id = str(get_col("ei_site_id") or "").strip()
            name = str(get_col("site_name") or "").strip()
            if not site_id or not name:
                skipped += 1
                errors.append(f"L{rnum}: 'site_id' ou 'name' manquant.")
                continue

            # Doublon → on ignore
            if Sites.objects.filter(site_id=site_id).exists():
                skipped += 1
                continue

            latitude = get_col("latitude")
            longitude = get_col("longitude")
            security_type = get_col("security_type")

            # Vendor par NOM
            vendor_obj = None
            vendor_name = get_col("vendor")
            if vendor_name:
                vendor_obj = Vendor.objects.filter(name__iexact=str(vendor_name).strip()).first()
                if not vendor_obj:
                    errors.append(f"L{rnum}: Vendor introuvable (name='{vendor_name}').")

            # RiskAssessment par NOM
            ra_obj = None
            ra_name = get_col("risk_assessment")
            if ra_name:
                ra_obj = RiskAssessment.objects.filter(name__iexact=str(ra_name).strip()).first()
                if not ra_obj:
                    errors.append(f"L{rnum}: RiskAssessment introuvable (name='{ra_name}').")

            # ZM par ID (UUID)
            zm_obj = None
            zm_id = get_col("zm")
            if zm_id:
                try:
                    zm_obj = User.objects.get(id=str(zm_id).strip())
                except User.DoesNotExist:
                    errors.append(f"L{rnum}: User (ZM) introuvable (id='{zm_id}').")
                except Exception as e:
                    errors.append(f"L{rnum}: ID ZM invalide ({zm_id}): {e}")

            Sites.objects.create(
                name=name,
                site_id=site_id,
                latitude=str(latitude).strip() if latitude is not None else None,
                longitude=str(longitude).strip() if longitude is not None else None,
                vendor=vendor_obj,
                risk_assessment=ra_obj,
                zm=zm_obj,
                security_type=str(security_type).strip() if security_type is not None else None
            )
            created += 1

        except Exception as e:
            errors.append(f"L{rnum}: {e}")

    return Response({"created": created, "skipped": skipped, "errors": errors}, status=200)

@extend_schema(
    tags=["sites"],
    summary="Lister tous les sites",
    description=(
        "Retourne la liste des sites, triée par `created_at` décroissant.\n\n"
        "**Filtres (query params)** :\n"
        "- `vendor` : UUID du Vendor (FK)\n"
        "- `risk_assessment` : UUID du RiskAssessment (FK)\n"
        "- `filter` : recherche texte (icontains) sur `name` et `site_id`"
    ),
    parameters=[
        OpenApiParameter(
            name="vendor",
            description="UUID du Vendor",
            required=False,
            type=OpenApiTypes.UUID,
            location=OpenApiParameter.QUERY,
        ),
        OpenApiParameter(
            name="risk_assessment",
            description="UUID du RiskAssessment",
            required=False,
            type=OpenApiTypes.UUID,
            location=OpenApiParameter.QUERY,
        ),
        OpenApiParameter(
            name="security_type",
            description="Filtre texte partiel (insensible à la casse) sur `security_type`",
            required=False,
            type=str,
            location=OpenApiParameter.QUERY,
        ),
        OpenApiParameter(
            name="filter",
            description="Filtre texte partiel (insensible à la casse) sur `name` et `site_id`",
            required=False,
            type=str,
            location=OpenApiParameter.QUERY,
        ),
    ],
    responses={
        200: OpenApiResponse(
            response=SitesSerializer(many=True),
            description="Liste des sites"
        ),
        500: OpenApiResponse(
            response=inline_serializer(
                name="SitesListServerError",
                fields={"message": serializers.CharField()}
            ),
            description="Erreur interne"
        ),
    },
    examples=[
        OpenApiExample(
            "Exemple de réponse 200 (partiel)",
            value=[
                {
                    "id": "0b8c2b3e-9f3b-4a1e-bf4b-1f3f1d2c9a11",
                    "name": "KASALA",
                    "site_id": "CDKN00001",
                    "latitude": "-4.325",
                    "longitude": "15.322",
                    "vendor": "59a0f6f9-4258-4fff-9a7c-0e103deffab1",
                    "risk_assessment": "3b4d0348-0c7c-4f18-9e0b-1b796b7b9d10",
                    "zm": "c7b0c8b1-1c0e-4803-8d6e-3a1b799c9f22",
                    "created_at": "2025-08-18T10:12:45Z"
                }
            ],
            response_only=True,
        )
    ],
)

@api_view(['GET'])
@authentication_classes([])
@permission_classes([])
def get_all_sites(request):
    """
    Récupère tous les sites enregistrés.
    """
    vendor_id = request.GET.get('vendor')
    risk_assessment = request.GET.get('risk_assessment')
    security_type_name = request.GET.get('security_type')
    search_filter = request.GET.get('filter', '').strip()

    filters = Q()
    if vendor_id:
        filters &= Q(vendor__id=vendor_id)
    if risk_assessment:
        filters &= Q(risk_assessment__id=risk_assessment)
    if security_type_name:
        filters &= Q(security_type=security_type_name)
    if search_filter:
        filters &= (
            Q(name__icontains=search_filter) |
            Q(site_id__icontains=search_filter) 
            # Q(si__icontains=search_filter) |
    #         Q(email__icontains=search_filter) |
    #         Q(telephone__icontains=search_filter) |
    #         Q(address__icontains=search_filter)
        )
    try:
        # Récupérer tous les sites
        # sites = Sites.objects.all()
        sites = Sites.objects.filter(filters).distinct().order_by('-created_at')

        


        # Sérialiser les sites
        serializer = SitesSerializer(sites, many=True)

        # Réponse avec pagination
        return Response(serializer.data)

    except Exception as e:
        
        print(e)
        return Response(
            {"message": str(e)},
            status=500
        )
    